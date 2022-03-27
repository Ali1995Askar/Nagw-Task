from collections import OrderedDict

import json
import openpyxl

from flask import jsonify, make_response
import docs


def get_url_from_formula(formula: str):
    str_list = formula.split('CONCATENATE')[1].split(',')
    url = str_list[0].split('"')[1] + str_list[1].split('"')[1]
    value = str_list[2].split('"')[1]
    return value, url


def get_books(request):
    """
        Get all books  method
    """
    books = []
    wb_obj = openpyxl.load_workbook(docs.FILE_NAME)
    sheet_obj = wb_obj.active

    for row in sheet_obj.iter_rows(min_row=2):
        book = OrderedDict()
        author = OrderedDict()
        country = OrderedDict()
        
        book['name'], book['url'] = get_url_from_formula(row[1].value)
        author['name'], author['url'] = get_url_from_formula(row[2].value)
        country['name'], country['url'] = get_url_from_formula(row[3].value)

        books_dict = OrderedDict()
        books_dict['id'] = row[0].row
        books_dict['order'] = row[0].value
        books_dict['book'] = book
        books_dict['author'] = author
        books_dict['country'] = country
        
        books.append(books_dict)
        
    wb_obj.close()
    
    return make_response(
        jsonify({'res': books}),
        200)


def create_book(request):
    """
        create book method
        request body should contain (order, author, book_name, country)
    """
    try:
        request_body = json.loads(request.data)
        order = request_body.get('order')
        book = docs.XLSX_FORMULA.format(
            request_body.get('book').get('url')[:len(request_body.get('book').get('url'))//2], 
            request_body.get('book').get('url')[len(request_body.get('book').get('url'))//2:], 
            request_body.get('book').get('name')
            )

        author = docs.XLSX_FORMULA.format(
            request_body.get('author').get('url')[:len(request_body.get('author').get('url'))//2], 
            request_body.get('author').get('url')[len(request_body.get('author').get('url'))//2:], 
            request_body.get('author').get('name')
            )

        country = docs.XLSX_FORMULA.format(
            request_body.get('country').get('url')[:len(request_body.get('country').get('url'))//2],
            request_body.get('country').get('url')[len(request_body.get('country').get('url'))//2:], 
            request_body.get('country').get('name') 
            
            )

        wb_obj = openpyxl.load_workbook(docs.FILE_NAME)
        sheet_obj = wb_obj.active

        row = [order, book, author, country]
        sheet_obj.append(row)
        wb_obj.save(docs.FILE_NAME)

        return make_response(
            jsonify({'res': 'Row Added succssfully'}),
                201
            )
    except Exception as e:
        return make_response(
                jsonify({'res': 'Something with your request body wrong (Bad Request)'}),
                400)


def update_book(request):
    """
        update book method
        request should contain (id)
        and new values of all or any (order, author, book_name, country)
    """
    try:
        request_body = json.loads(request.data)
        wb_obj = openpyxl.load_workbook(docs.FILE_NAME)
        sheet_obj = wb_obj.active
        id = request_body.get('id')
        order = request_body.get('order') if request_body.get('order') else None
        book = request_body.get('book') if request_body.get('book') else None
        author = request_body.get('author') if request_body.get('author') else None
        country = request_body.get('country') if request_body.get('country') else None
        row_to_update = sheet_obj[id]
        
        if order:
            row_to_update[0].value = order if order else row_to_update[0].value

        if book:
            row_to_update[1].value = docs.XLSX_FORMULA.format(
                book.get('url')[:len(book.get('url'))//2], 
                book.get('url')[len(book.get('url'))//2:], 
                book.get('name'),  
            )
           
        if book:
            row_to_update[2].value = docs.XLSX_FORMULA.format(
                author.get('url')[:len(author.get('url'))//2], 
                author.get('url')[len(author.get('url'))//2:], 
                author.get('name'),  
            )

        if country:
            row_to_update[3].value = docs.XLSX_FORMULA.format(
                country.get('url')[:len(country.get('url'))//2], 
                country.get('url')[len(country.get('url'))//2:], 
                country.get('name'),  
            )

        wb_obj.save(docs.FILE_NAME)

        return make_response(
        jsonify({'res': 'Row Updated Successfully !!'}),
            200
        )

    except:
        return make_response(
            jsonify({'res': 'Something with your request body wrong (Bad Request)'}),
            400)


def delete_book(request):
    """
        delete book method
        request should contain (id)
    """
    try:
        request_body = json.loads(request.data)
        id = request_body.get('id')
        wb_obj = openpyxl.load_workbook(docs.FILE_NAME)
        sheet_obj = wb_obj.active

        sheet_obj.delete_rows(idx=id)
        wb_obj.save(docs.FILE_NAME)

        return make_response(
            jsonify({'res': 'Row Deleted Successfully'}),
            200)
    except:
        return make_response(
            jsonify({'res': 'Something with your request body wrong (Bad Request)'}),
            400)


def api_docs(request):
    """
        simple Api docs
    """
    return make_response(
        jsonify({

            '***GET-BOOKS***': docs.GET_BOOKS_REQUEST_STRUCTURE,
            '***CREATE-BOOK***': docs.CREATE_BOOK_REQUEST_STRUCTURE,
            '***UPDATE-BOOK***': docs.UPDATE_BOOK_REQUEST_STRUCTURE,
            '***DELETE-BOOK***': docs.DELETE_BOOK_REQUEST_STRUCTURE,
        }),
        200
    )






